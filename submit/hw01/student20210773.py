import pandas as pd
from openpyxl import load_workbook

# 워크북 불러오기
workbook = load_workbook("23_02_bigdata\submit\hw01\student.xlsx")
sheet = workbook.active

sheet.delete_cols(7)
sheet.delete_cols(8)
sheet.cell(row=1, column=7, value="total")
sheet.cell(row=1, column=8, value="grade")

# 데이터 추출
data = list(sheet.iter_rows(values_only=True))
header = data[0]
scores = [list(row) for row in data[1:]]

# 총점 및 학점 계산
for score in scores:
    midterm, final, hw, attendance = score[2:6]
    total = midterm * 0.3 + final * 0.35 + hw * 0.34 + attendance * 0.01
    score.append(total)

# 총점을 기준으로 정렬
scores.sort(key=lambda x: x[7] if x[7] is not None else 0, reverse=True)

# 학점 부여
num_students = len(scores)
grades = ["A+"] * (num_students // 6) + ["A"] * (num_students // 3) + ["B+"] * (num_students // 6) + \
         ["B"] * (num_students // 6) + ["C+"] * (num_students // 6) + ["C"] * (num_students - (num_students // 6) * 5)

# 총점이 40 미만인 학생에게 F 학점 부여
for i in range(num_students):
    if scores[i][7] is not None and scores[i][7] < 40:
        grades[i] = "F"
    scores[i].append(grades[i] if grades[i] is not None else "")

# id 값을 기준으로 학생 데이터 재배열
sorted_scores = sorted(scores, key=lambda x: int(x[0]))
sorted_grades = [grades[scores.index(student)] for student in sorted_scores]

# 워크시트에 쓰기
for i, row in enumerate(sorted_scores):
    for j, value in enumerate(row):
        try:
            sheet.cell(row=i + 2, column=j - 1, value=value)  # 데이터가 2행부터 시작한다고 가정
        except ValueError:
            pass


# 워크북 저장
workbook.save("23_02_bigdata\submit\hw01\student_with_grades.xlsx")
